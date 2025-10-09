from math import ceil
import uuid, json, base64, requests
import requests  # type: ignore
from django.shortcuts import redirect, resolve_url
from io import BytesIO 
from decimal import Decimal 
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings  # type: ignore
from django.contrib import messages  # type: ignore
from django.contrib.auth import login, logout, get_user_model
from django.contrib.auth.decorators import (
    login_required,
    user_passes_test
)
from django.contrib.auth.forms import AuthenticationForm
from django.core.exceptions import ValidationError
from django.core.mail import EmailMessage, EmailMultiAlternatives
from django.db import transaction
from django.db.models import Q, Case, When, IntegerField, Count
from django.http import (
    HttpResponse,
    JsonResponse,
    HttpResponseBadRequest,
)

from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.encoding import force_bytes, force_str
from django.utils.html import strip_tags
from django.utils.http import (
    urlsafe_base64_encode,
    urlsafe_base64_decode
)
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import (
    require_http_methods,
    require_POST
)
from django.views.generic import (
    CreateView,
    ListView,
    DetailView,
    UpdateView,
    DeleteView,
)

from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font, Alignment  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

from .forms import (
    UserRegisterForm,
    CourseDetailsForm,
    PersonalInfoForm,
    EducationalInfoForm,
    RejectReasonForm,
    UserContactForm,
    PaymentDetailForm,
)
from .models import (
    CourseDetails,
    PersonalInfo,
    EducationalInfo,
    Application,
    PaymentDetail,
    Notification,
)

#impors for custom passwordchange 
from django.contrib.auth.views import PasswordChangeView
from .tokens import account_activation_token  


User = get_user_model()


# -----------------------------
# Helper: Admin check
# -----------------------------
def _is_admin(user):
    return getattr(user, "is_admin", False) or user.is_staff


# -----------------------------
# Helper: send activation email
# -----------------------------
def send_activation_email(user, request):
    uidb64 = urlsafe_base64_encode(force_bytes(user.pk))
    token = account_activation_token.make_token(user)
    activation_url = request.build_absolute_uri(
        reverse(
            "activate",
            kwargs={"uidb64": uidb64, "token": token}
        )
    )
    subject = "Verify your email"
    body = (
        f"Hi {user.username},\n\n"
        f"Please click the link below to verify your email and "
        f"activate your account:\n\n"
        f"{activation_url}\n\n"
        f"If you didn't sign up, you can ignore this message."
    )
    EmailMessage(
        subject,
        body,
        settings.DEFAULT_FROM_EMAIL,
        [user.email]
    ).send()

#------------------------------------
#Helper: Notification Object Creation
#------------------------------------ 
def create_notification(user, title, message):
    notification = Notification.objects.create(
        user=user,
        title=title,
        message=message,
    )


# -----------------------------
# Register View
# -----------------------------
def register(request):
    if request.method == "POST":
        form = UserRegisterForm(request.POST, request)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False
            user.save()
            
            # Create notification
            create_notification(
                user=user,
                title="Welcome to Online Enrollment System!",
                message="Your account has been created successfully.",
            )
            send_activation_email(user, request)
            # messages.success(
            #     request,
            #     "Registration successful! Check your email to verify "
            #     "your account.",
            # )
            return redirect("login_page")
    else:
        form = UserRegisterForm()
    return render(
        request,
        "registration/register.html",
        {"form": form}
    )


# -----------------------------
# Activate account from email link
# -----------------------------
# def activate(request, uidb64, token):
#     try:
#         uid = force_str(urlsafe_base64_decode(uidb64))
#         user = User.objects.get(pk=uid)
#     except (TypeError, ValueError, OverflowError, User.DoesNotExist):
#         user = None

#     if user and account_activation_token.check_token(user, token):
#         user.is_active = True
#         user.save(update_fields=["is_active"])
#         messages.success(
#             request,
#             "Your email has been verified. You can now log in."
#         )
#         return redirect("login_page")

#     messages.error(
#         request,
#         "Activation link is invalid or has expired."
#     )
#     return redirect("register")


def activate(request, uidb64, token):
    user = None
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except Exception as e:
        # Optional: log this in dev
        print("Activation decode/get error:", e)

    login_url = resolve_url("login_page")  # or "admissionapp:login_page" if namespaced
    register_url = resolve_url("register")

    if user and account_activation_token.check_token(user, token):
        if not user.is_active:
            user.is_active = True
            user.save(update_fields=["is_active"])
        messages.success(request, "Your email has been verified. You can now log in.")
        return redirect(login_url)

    messages.error(request, "Activation link is invalid or has expired.")
    return redirect(register_url)


# -----------------------------
# Login View
# -----------------------------
def login_page(request):
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()

            # Block user if email is not verified
            if not user.is_active:
                messages.error(
                    request,
                    "Please, verify your email before logging in."
                )
                return redirect("login_page")

            login(request, user)

            # Role-based dashboard
            if getattr(user, "is_admin", True):
                dashboard = "admin_dashboard"
            else:
                dashboard = "student_dashboard"
            return redirect(dashboard)
    else:
        form = AuthenticationForm()
    return render(
        request,
        "registration/login_page.html",
        {"form": form}
    )


# -----------------------------
# Custom Redirect
# -----------------------------
def custom_redirect_url(request):
    if request.user.is_admin:
        return redirect("admin_dashboard")
    return redirect("student_dashboard")


# -----------------------------
# Student Dashboard helpers
# -----------------------------
def repeat_to_fill(seq, per_row=3, min_rows=1, max_rows=4):
    seq = list(seq)
    if not seq:
        return []
    target_len = per_row * ceil(len(seq) / per_row)
    target_len = max(target_len, per_row * min_rows)
    target_len = min(target_len, per_row * max_rows)
    repeats = (target_len + len(seq) - 1) // len(seq)
    return (seq * repeats)[:target_len]


# -----------------------------
# Student Dashboard
# -----------------------------
def student_dashboard(request):
    plus2_courses = CourseDetails.objects.filter(
        degree="Plus2"
    ).order_by("course_name")
    bachelor_courses = CourseDetails.objects.filter(
        degree="Bachelor"
    ).order_by("course_name")
    master_courses = CourseDetails.objects.filter(
        degree="Master"
    ).order_by("course_name")

    # Single query to get all counts
    course_stats = CourseDetails.objects.aggregate(
        total_courses=Count("id"),
        master_course_count=Count(
            "id",
            filter=Q(degree="Master")
        ),
        bachelor_course_count=Count(
            "id",
            filter=Q(degree="Bachelor")
        ),
        plus2_course_count=Count(
            "id",
            filter=Q(degree="Plus2")
        ),
    )

    context = {
        "total_courses": course_stats["total_courses"],
        "master_course_count": course_stats["master_course_count"],
        "bachelor_course_count": (
            course_stats["bachelor_course_count"]
        ),
        "plus2_course_count": course_stats["plus2_course_count"],
        "plus2_courses": plus2_courses,
        "bachelor_courses": bachelor_courses,
        "master_courses": master_courses,
        # Repeated lists for the grid
        "plus2_courses_repeat": repeat_to_fill(
            plus2_courses,
            per_row=3,
            min_rows=1,
            max_rows=4
        ),
        "bachelor_courses_repeat": repeat_to_fill(
            bachelor_courses,
            per_row=3,
            min_rows=1,
            max_rows=4
        ),
        "master_courses_repeat": repeat_to_fill(
            master_courses,
            per_row=3,
            min_rows=1,
            max_rows=4
        ),
    }
    return render(
        request,
        "student/student_dashboard.html",
        context
    )


# -----------------------------
# Admin Dashboard
# -----------------------------
def admin_dashboard(request):
    # Course related statistics
    course_stats = CourseDetails.objects.aggregate(
        total=Count("id"),
        master=Count("id", filter=Q(degree="Master")),
        bachelor=Count("id", filter=Q(degree="Bachelor")),
        plus2=Count("id", filter=Q(degree="Plus2")),
    )

    # Applications related statistics
    stats = Application.objects.aggregate(
        total=Count("id"),
        approved=Count(
            "id",
            filter=Q(application_status="approved")
        ),
        rejected=Count(
            "id",
            filter=Q(application_status="rejected")
        ),
        pending=Count(
            "id",
            filter=Q(application_status="pending")
        ),
    )

    context = {
        "total_course": course_stats["total"],
        "total_master_course": course_stats["master"],
        "total_bachelor_course": course_stats["bachelor"],
        "total_plus2_course": course_stats["plus2"],
        # Application stats
        "total_applications": stats["total"],
        "approved_applications": stats["approved"],
        "rejected_applications": stats["rejected"],
        "pending_applications": stats["pending"],
    }
    return render(request, "admin/admin_dashboard.html", context)


# -----------------------------
# Logout
# -----------------------------
def log_out(request):
    logout(request)
    return redirect("login_page")


# -----------------------------
# Add Course (CreateView)
# -----------------------------
class AddCourseView(CreateView):
    model = CourseDetails
    form_class = CourseDetailsForm
    template_name = "admin/add_course.html"
    success_url = reverse_lazy("course_list")


# -----------------------------
# Course List (ListView)
# -----------------------------
class CourseListView(ListView):
    model = CourseDetails
    template_name = "admin/course_list.html"
    context_object_name = "courses"

    def get_queryset(self):
        degree_order = Case(
            When(degree="Plus2", then=0),
            When(degree="Bachelor", then=1),
            When(degree="Master", then=2),
            default=3,
            output_field=IntegerField(),
        )
        return (
            CourseDetails.objects
            .annotate(_degree_order=degree_order)
            .order_by("_degree_order", "course_name")
        )


# -----------------------------
# Course Update (UpdateView)
# -----------------------------
class CourseUpdateView(UpdateView):
    model = CourseDetails
    form_class = CourseDetailsForm
    template_name = "admin/course_update.html"
    success_url = reverse_lazy("course_list")


# -----------------------------
# Course Delete (DeleteView)
# -----------------------------
class CourseDeleteView(DeleteView):
    model = CourseDetails
    success_url = reverse_lazy("course_list")


# -----------------------------
# Course Detail (DetailView)
# -----------------------------
class CourseDetailView(DetailView):
    model = CourseDetails
    template_name = "admin/course_detail.html"
    context_object_name = "courses"


# -----------------------------
# Personal Info (Create)
# -----------------------------
@login_required
def PersonalInfo_view(request):
    # Check if the user already has a personal info record
    personalinfo = PersonalInfo.objects.filter(
        user=request.user
    ).first()

    # If exists, redirect to detail
    if personalinfo:
        return redirect("personalinfo_detail", pk=personalinfo.pk)

    # If no record, let the form create a new record
    if request.method == "POST":
        form = PersonalInfoForm(request.POST, request.FILES)
        if form.is_valid():
            personalinfo = form.save(commit=False)
            personalinfo.user = request.user
            personalinfo.save()
            
            # Create notification
            create_notification(
                user=request.user,
                title="Personal Information Saved",
                message="Your personal details have been successfully submitted and saved.",
            )
            messages.success(
                request,
                "Personal Details submitted successfully"
            )
            return redirect("educational_info")
    else:
        form = PersonalInfoForm()
    return render(
        request,
        "student/personal_info.html",
        {"form": form}
    )


# -----------------------------
# Educational Info (Create)
# -----------------------------
@login_required
def EducationalInfo_view(request):
    educationinfo = EducationalInfo.objects.filter(
        user=request.user
    ).first()

    # Display the information on the same page if object exists.
    if educationinfo:
        base_qs = (
            EducationalInfo.objects
            .filter(user=request.user)
            .select_related("user")
            .order_by("passed_year")
        )
        # (kept as in original comment)

    if request.method == "POST":
        try:
            edu = EducationalInfo.objects.create(
                user=request.user,
                level=request.POST.get("level", "SEE"),
                faculty=request.POST.get("faculty", "Commerce"),
                course_name=request.POST.get("course_name"),
                university_name=request.POST.get(
                    "university_name",
                    "Tribhuvan University"
                ),
                college_name=request.POST.get("college_name"),
                passed_year=request.POST.get("passed_year"),
                grade_percent=request.POST.get("grade_percent"),
                upload_transcript1=request.FILES.get(
                    "upload_transcript1"
                ),
                upload_transcript2=request.FILES.get(
                    "upload_transcript2"
                ),
                upload_character=request.FILES.get(
                    "upload_character"
                ),
                upload_license=request.FILES.get("upload_license"),
                upload_other=request.FILES.get("upload_other"),
                upload_other1=request.FILES.get("upload_other1"),
                created_at=timezone.now(),
                updated_at=timezone.now(),
            )

            # Run full model validation (validators for year, % etc.)
            edu.full_clean()
            edu.save() 
            
            # Create notification
            create_notification(
                user=request.user,
                title="Educational Information Saved",
                message="Your educational details have been successfully submitted and saved.",
            )

            messages.success(
                request,
                "✅ Educational details saved successfully!"
            )
            return redirect("education_list")

        except ValidationError as e:
            for field, errors in e.message_dict.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
        except Exception as e:
            messages.error(request, f"An error occurred: {e}")

    return render(request, "student/educational_info.html")


#-----------------------------
#Notification View 
#-----------------------------
@login_required
def notification_list(request):
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at') 
    context = {
        'notifications':notifications,
    }

    return render(request, 'student/notification_list.html', context)
# -----------------------------
# Profile
# -----------------------------
@login_required
def profile(request):
    user = request.user
    personalinfo = getattr(user, "personal_info", None)
    context = {
        "user": user,
        "personalinfo": personalinfo,
    }
    return render(request, "student/profile.html", context)


# -----------------------------
# PersonalInfo Detail
# -----------------------------
@login_required
def PersonalInfo_Detail(request, pk):
    # Check if any application has been already approved
    application_approved = Application.objects.filter(
        user=request.user,
        application_status="approved"
    ).exists()

    per_info = get_object_or_404(PersonalInfo, pk=pk)

    # Prevent users from seeing other's detail
    if per_info.user != request.user:
        messages.error(
            request,
            "You are not allowed to view this information"
        )
        return redirect(
            "personalinfo_detail",
            pk=request.user.per_info.pk
        )

    context = {
        "per_info": per_info,
        "application_approved": application_approved,
    }
    return render(
        request,
        "student/personalinfo_detail.html",
        context
    )


# -----------------------------
# EducationalInfo Detail
# -----------------------------
@login_required
def EducationalInfo_Detail(request, pk):
    educationalinfo = get_object_or_404(EducationalInfo, pk=pk)

    # Prevent users from seeing other's detail
    if educationalinfo.user != request.user:
        messages.error(
            request,
            "You are not allowed to view this information"
        )
        return redirect(
            "educationalinfo_detail",
            pk=request.user.educationalinfo.pk
        )

    context = {"edu": educationalinfo}
    return render(
        request,
        "student/educationalinfo_detail.html",
        context
    )


# -----------------------------
# EducationalInfo List
# -----------------------------
def education_list_view(request):
    application_approved = Application.objects.filter(
        user=request.user,
        application_status="approved"
    ).exists()
    educationinfo = EducationalInfo.objects.filter(
        user=request.user
    ).first()

    if not educationinfo:
        return redirect("educational_info")

    base_qs = (
        EducationalInfo.objects
        .filter(user=request.user)
        .select_related("user")
        .order_by("passed_year")
    )
    education_count = base_qs.count()

    context = {
        "see_details": base_qs.filter(level="SEE"),
        "plus2_details": base_qs.filter(level="Plus2"),
        "bachelor_details": base_qs.filter(level="Bachelor"),
        "master_details": base_qs.filter(level="Master"),
        "education_count": education_count,
        "application_approved": application_approved,
    }
    return render(
        request,
        "student/education_list.html",
        context
    )


# -----------------------------
# Edit PersonalInfo
# -----------------------------
def edit_personal_info(request, pk):
    student = get_object_or_404(
        PersonalInfo,
        pk=pk,
        user=request.user
    )

    if request.method == "POST":
        form = PersonalInfoForm(
            request.POST,
            request.FILES,
            instance=student
        )
        if form.is_valid():
            form.save()
            
            # Create notification
            create_notification(
                user=request.user,
                title="Personal Detail Updated",
                message="Your personal details have been successfully submitted and saved.",
            )
            return redirect("student_dashboard")
    else:
        form = PersonalInfoForm(instance=student)
    return render(
        request,
        "student/edit_personalinfo.html",
        {"form": form}
    )


# -----------------------------
# Edit EducationalInfo
# -----------------------------
def edit_educational_info(request, pk):
    edu_info = get_object_or_404(EducationalInfo, pk=pk)
    if request.method == "POST":
        form = EducationalInfoForm(
            request.POST,
            request.FILES,
            instance=edu_info
        )
        if form.is_valid():
            form.save()
            
            # Create notification
            create_notification(
                user=request.user,
                title="Educational Information Updated",
                message="Your educational details have been successfully updated and saved.",
            )
            return redirect("education_list")
    else:
        form = EducationalInfoForm(instance=edu_info)
    return render(
        request,
        "student/edit_educationalinfo.html",
        {"form": form}
    )


def course_application_list(request):
    personal_info = PersonalInfo.objects.filter(
        user=request.user
    ).exists()
    
    edu_info = EducationalInfo.objects.filter(
        user=request.user
    ).exists()
     
    payment_info = PaymentDetail.objects.filter(
        user=request.user).exists()
     
    applicants = Application.objects.select_related(
        "user",
        "course"
    ).all()

    context = {
        "per_info": personal_info,
        "edu_info": edu_info,
        "applicants": applicants,
        'payment_info':payment_info,
    }
    return render(
        request,
        "admin/course_applicant_list.html",
        context
    )


# --------------------------------------------
# Course Applicant Detail:Admin Function
# ---------------------------------------------
def course_applicant_detail(request, pk):
    # Get the applicants instance
    applicant = get_object_or_404(Application, pk=pk) 
    

    # Get applicant related personal info and educational info
    personal_info = PersonalInfo.objects.filter(
        user=applicant.user
    ).first()
    
    edu_info = EducationalInfo.objects.filter(
        user=applicant.user
    ) 
    
    document_list = []
    if edu_info:
        for edu in edu_info: 
            docs = [
                ("Transcript 1", edu.upload_transcript1),
                ("Transcript 2", edu.upload_transcript2),
                ("Character Certificate", edu.upload_character),
                ("License", edu.upload_license),
                ("Other Document", edu.upload_other),
                ("Other Document 1", edu.upload_other1),
            ]
            document_list.append(
                {
                    "edu":edu,
                    "documents": [doc for doc in docs if doc[1]] 
                }
            )
    
    payment_info = PaymentDetail.objects.filter(application=applicant).select_related('application__course').order_by('-payment_date').first()

    context = {
        "per_info": personal_info,
        "edu_info": edu_info,
        "applicant": applicant,
        "payment_info": payment_info,
        "document_info":document_list,
    }
    return render(
        request,
        "admin/course_applicant_detail.html",
        context
    )


# -----------------------------
# Select Course by Student
# -----------------------------
def select_course(request, pk):
    course_info = get_object_or_404(CourseDetails, pk=pk) 
    
    #check if pending application exist 
    application_exists = Application.objects.filter(
        user=request.user, application_status='pending'
    ).exists()
    
    has_personal_info = PersonalInfo.objects.filter(
        user=request.user
    ).exists()
    has_edu_info = EducationalInfo.objects.filter(
        user=request.user
    ).exists()

    context = {
        "course_info": course_info,
        "has_personal_info": has_personal_info,
        "has_edu_info": has_edu_info,
        "application_exists": application_exists,
    }
    return render(request, "student/select_course.html", context)


# -----------------------------
# Apply Course (Online Application)
# -----------------------------
def apply_course(request, pk):
    course = get_object_or_404(CourseDetails, pk=pk)
    personal_info = PersonalInfo.objects.filter(
        user=request.user
    ).exists()
    edu_info = EducationalInfo.objects.filter(
        user=request.user
    ).exists()

    # First, check personalinfo and educationinfo object
    if not personal_info or not edu_info:
        messages.info(
            request,
            "Insert Personal Info and Educational info "
            "before apply."
        )
        return redirect("student_dashboard")

    # Check for one active application
    already_has_application = (
        Application.objects
        .filter(user=request.user)
        .exclude(application_status="rejected")
        .exists()
    )
    if already_has_application:
        messages.error(
            request,
            "You can't apply for more than one course."
        )
        return redirect("student_dashboard")

    application, created = Application.objects.get_or_create(
        user=request.user,
        course=course
    )

    if created:
        # Create notification
        create_notification(
            user=request.user,
            title="Applied Course",
            message="Your Application has been successfully submitted and saved.",
        )
        
        # Create notification
        create_notification(
            user=request.user,
            title="Course Application",
            message="Your application has been successfully submitted and saved.",
        )
        
        messages.success(
            request,
            "Application Submitted Successfully."
        )
    else:
        messages.info(
            request,
            "You already applied for this course."
        )
    return redirect("application_list")


# -----------------------------
# Student Application List
# -----------------------------
def application_list(request):
    applications = (
        Application.objects
        .filter(user=request.user)
        .select_related("course")
        .prefetch_related("payment")
        .order_by("submitted_at")
    )
    
    payment_info = PaymentDetail.objects.filter(
        user=request.user, status="COMPLETE").select_related("application__course").first()
    
    
    context = {
        "applications": applications,
        "payment_info":payment_info,
    }
    
    return render(
        request,
        "student/application_list.html",context)


# -----------------------------
# Course Approval/Rejection: Admin Side
# -----------------------------
@require_POST
@login_required
@user_passes_test(_is_admin)
def approval_rejection(request, pk):
    action = request.POST.get("action")

    with transaction.atomic():
        # Lock the application row
        application = get_object_or_404(
            Application.objects
            .select_for_update()
            .select_related("course", "user"),
            pk=pk,
        )
        # Lock the related course row to prevent seat races
        course = CourseDetails.objects.select_for_update().get(
            pk=application.course_id
        )

        if action == "approve":
            if application.application_status == "approved":
                messages.info(
                    request,
                    "This application is already approved."
                )
                return redirect("course_application_list")

            # Capacity check
            if course.seats_filled >= course.total_seats:
                messages.error(
                    request,
                    "No seats available for this course."
                )
                return redirect("course_application_list")

            # Reserve seat first
            course.seats_filled += 1
            course.save(update_fields=["seats_filled"])

            # Mark application approved
            application.application_status = "approved"
            application.approved_rejected_date = timezone.now()
            application.reason_to_reject = ""
            application.save(
                update_fields=[
                    "application_status",
                    "approved_rejected_date",
                    "reason_to_reject",
                ]
            )
            
            # Create notification
            create_notification(
                user=application.user, 
                title="Application Approved",
                message="Congratulations! Your application has been approved.")
            
            # Email (only if recipient exists)
            recipient = application.user.email
            if recipient:
                context = {
                    "application": application,
                    "applicant_name": (
                        application.user.get_full_name() or
                        application.user.username
                    ),
                    "course_name": course.course_name,
                    "approved_rejected_date": (
                        application.approved_rejected_date
                    ),
                    "status": "approved",
                    "site_name": "Online Admission System",
                    "site_team_name": "Admissions Team",
                    "support_email": "support@example.com",
                }
                html_body = render_to_string(
                    "emails/email_to_send.html",
                    context
                )
                text_body = strip_tags(html_body)

                email = EmailMultiAlternatives(
                    subject="Application Approval",
                    body=text_body,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[recipient],
                )
                email.attach_alternative(html_body, "text/html")
                email.send(fail_silently=True)

            messages.success(
                request,
                "Application approved and email sent!"
            )
            return redirect("course_application_list")

        elif action == "reject":
            # Optional: forbid rejecting already-approved app
            if application.application_status == "approved":
                messages.error(
                    request,
                    "Approved application cannot be rejected."
                )
                return redirect("course_application_list")

            application.application_status = "rejected"
            application.approved_rejected_date = timezone.now()
            application.save(
                update_fields=[
                    "application_status",
                    "approved_rejected_date"
                ]
            )
            return redirect("reason_to_reject", pk=application.pk)
        else:
            messages.error(request, "Invalid action.")
            return redirect("course_application_list")


# -----------------------------
# Reason to Reject: Admin
# -----------------------------
@require_http_methods(["GET", "POST"])
@login_required
@user_passes_test(_is_admin)
def reason_to_rejection(request, pk):
    application = get_object_or_404(
        Application.objects.select_related("course", "user"),
        pk=pk,
    )

    if request.method == "POST":
        form = RejectReasonForm(request.POST, instance=application)
        if form.is_valid():
            app = form.save(commit=False)

            # Ensure status & timestamp are set
            if app.application_status != "rejected":
                app.application_status = "rejected"
            if not app.approved_rejected_date:
                app.approved_rejected_date = timezone.now()

            app.save()
            
            # Create notification
            create_notification(
                user=request.user,
                title="Applicaton Rejection",
                message="Your application has been Rejected and saved.",
            )

            # Build email context from updated instance (app)
            context = {
                "application": app,
                "applicant_name": (
                    app.user.get_full_name() or
                    app.user.username
                ),
                "course_name": app.course.course_name,
                "approved_rejected_date": (
                    app.approved_rejected_date
                ),
                "status": "rejected",
                "site_name": "Online Admission System",
                "site_team_name": "Admissions Team",
                "support_email": "support@example.com",
            }

            html_body = render_to_string(
                "emails/email_to_send.html",
                context
            )
            text_body = strip_tags(html_body)

            recipient = app.user.email
            if recipient:
                email = EmailMultiAlternatives(
                    subject="Application Rejection",
                    body=text_body,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[recipient],
                )
                email.attach_alternative(html_body, "text/html")
                email.send(fail_silently=True)
                messages.success(
                    request,
                    "Application rejected and email sent "
                    "successfully!"
                )
                
                
            else:
                messages.warning(
                    request,
                    "Application rejected, but no email found "
                    "for user."
                )

            return redirect("course_application_list")
    else:
        form = RejectReasonForm(instance=application)

    return render(
        request,
        "admin/reason_to_reject.html",
        {"form": form, "application": application},
    )


# -----------------------------
# Re-submission for Rejected Application
# -----------------------------
def re_submit_application(request, pk):
    application = get_object_or_404(
        Application,
        pk=pk,
        user=request.user
    )

    if application.application_status == "rejected":
        application.application_status = "re-submit"
        application.save()
        
        # Create notification
        create_notification(
            user=request.user,
            title="Application Re-Submission",
            message="Your application has been successfully Re-Submited.",
        )
        messages.success(
            request,
            "Application re-submitted successfully."
        )
    return redirect("application_list")


# -----------------------------
# Reports (Admin)
# -----------------------------
@login_required
@user_passes_test(_is_admin)
def reports(request):
    return render(request, "admin/reports.html")


@login_required
@user_passes_test(_is_admin)
def total_applications_report(request):
    qs = Application.objects.all()
    return render(
        request,
        "admin/reports/total_applications_report.html",
        {"qs": qs}
    )


@login_required
@user_passes_test(_is_admin)
def total_approved_report(request):
    qs = Application.objects.filter(application_status="approved")
    return render(
        request,
        "admin/reports/total_approved_report.html",
        {"qs": qs}
    )


@login_required
@user_passes_test(_is_admin)
def total_pending_report(request):
    qs = Application.objects.filter(application_status="pending")
    return render(
        request,
        "admin/reports/total_pending_report.html",
        {"qs": qs}
    )


@login_required
@user_passes_test(_is_admin)
def total_rejected_report(request):
    qs = Application.objects.filter(
        application_status="rejected"
    )
    return render(
        request,
        "admin/reports/total_rejected_report.html",
        {"qs": qs}
    )


# -----------------------------
# Export: Total Applications (Excel)
# -----------------------------
@login_required
@user_passes_test(_is_admin)
def export_total_applications(request):
    """
    Export applications to an Excel file.
    Optional GET params:
      - status=pending|approved|rejected|re-submit
      - start=YYYY-MM-DD (filter by submitted_at >= start)
      - end=YYYY-MM-DD   (filter by submitted_at <= end, inclusive)
    """
    qs = Application.objects.select_related(
        "user",
        "course"
    ).order_by("-submitted_at")

    status = request.GET.get("status")
    if status:
        qs = qs.filter(application_status=status)

    start_str = request.GET.get("start")
    end_str = request.GET.get("end")
    if start_str:
        start = parse_date(start_str)
        if start:
            qs = qs.filter(submitted_at__date__gte=start)
    if end_str:
        end = parse_date(end_str)
        if end:
            qs = qs.filter(submitted_at__date__lte=end)

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    headers = [
        "SN",
        "Name",
        "Mobile",
        "Application No.",
        "Applied Course",
        "Applied Degree",
        "Application Status",
        "Submitted Date",
        "Approved/Rejected Date",
    ]
    ws.append(headers)

    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for idx, app in enumerate(qs, start=1):
        full_name = (
            getattr(app.user, "full_name", "").strip() or
            app.user.get_username()
        )
        submitted_str = app.submitted_at.strftime(
            "%Y-%m-%d %H:%M"
        )
        approved_rejected_date_str = (
            app.approved_rejected_date.strftime("%Y-%m-%d %H:%M")
        )
        ws.append(
            [
                idx,
                full_name,
                app.user.mobile or "",
                app.application_no,
                app.course.course_name,
                app.course.degree,
                app.application_status.capitalize(),
                submitted_str,
                approved_rejected_date_str,
            ]
        )

    for column_cells in ws.columns:
        max_len = 0
        col = column_cells[0].column
        for c in column_cells:
            val = str(c.value) if c.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col)].width = min(
            max_len + 2,
            50
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = "total_applications.xlsx"
    if status:
        filename = f"applications_{status}.xlsx"

    content_type = (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )
    resp = HttpResponse(buffer.getvalue(), content_type=content_type)
    resp["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )
    return resp


# -----------------------------
# Export: Approved Applications (Excel)
# -----------------------------
@login_required
@user_passes_test(_is_admin)
def export_approved_applications(request):
    """Export only approved applications into an Excel file."""
    qs = (
        Application.objects
        .select_related("user", "course")
        .filter(application_status="approved")
        .order_by("-submitted_at")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Approved Applications"

    headers = [
        "SN",
        "Name",
        "Mobile",
        "Application No.",
        "Applied Course",
        "Applied Degree",
        "Application Status",
        "Approved On",
    ]
    ws.append(headers)

    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for idx, app in enumerate(qs, start=1):
        full_name = (
            getattr(app.user, "full_name", "").strip() or
            app.user.get_username()
        )
        approved_str = (
            app.approved_rejected_date.strftime("%Y-%m-%d %H:%M")
            if app.approved_rejected_date
            else ""
        )
        ws.append(
            [
                idx,
                full_name,
                app.user.mobile or "",
                app.application_no,
                app.course.course_name,
                app.course.degree,
                app.application_status,
                approved_str,
            ]
        )

    for column_cells in ws.columns:
        max_len = 0
        col = column_cells[0].column
        for c in column_cells:
            val = str(c.value) if c.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col)].width = min(
            max_len + 2,
            50
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    content_type = (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )
    resp = HttpResponse(buffer.getvalue(), content_type=content_type)
    resp["Content-Disposition"] = (
        'attachment; filename="approved_applications.xlsx"'
    )
    return resp


# -----------------------------
# Export: Rejected Applications (Excel)
# -----------------------------
@login_required
@user_passes_test(_is_admin)
def export_rejected_applications(request):
    """Export only rejected applications into an Excel file."""
    qs = (
        Application.objects
        .select_related("user", "course")
        .filter(application_status="rejected")
        .order_by("-submitted_at")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Rejected Applications"

    headers = [
        "SN",
        "Name",
        "Mobile",
        "Application No.",
        "Applied Course",
        "Applied Degree",
        "Application Status",
        "Rejection Date",
        "Reason to Reject",
    ]
    ws.append(headers)

    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for idx, app in enumerate(qs, start=1):
        full_name = (
            getattr(app.user, "full_name", "").strip() or
            app.user.get_username()
        )
        rejected_str = (
            app.approved_rejected_date.strftime("%Y-%m-%d %H:%M")
            if app.approved_rejected_date
            else ""
        )
        ws.append(
            [
                idx,
                full_name,
                app.user.mobile or "",
                app.application_no,
                app.course.course_name,
                app.course.degree,
                app.application_status,
                rejected_str,
                app.reason_to_reject or "",
            ]
        )

    for column_cells in ws.columns:
        max_len = 0
        col = column_cells[0].column
        for c in column_cells:
            val = str(c.value) if c.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col)].width = min(
            max_len + 2,
            50
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    content_type = (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )
    resp = HttpResponse(buffer.getvalue(), content_type=content_type)
    resp["Content-Disposition"] = (
        'attachment; filename="rejected_applications.xlsx"'
    )
    return resp


#-----------------------------
#Export Pending Applications 
#-----------------------------

@login_required
@user_passes_test(_is_admin)
def export_pending_applications(request):
    """Export only pending applications into an Excel file."""
    qs = (
        Application.objects
        .select_related("user", "course")
        .filter(application_status="pending")
        .order_by("-submitted_at")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Pending Applications"

    headers = [
        "SN",
        "Name",
        "Mobile",
        "Application No.",
        "Applied Course",
        "Applied Degree",
        "Application Status",
        "Submitted Date",
    ]
    ws.append(headers)

    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, app in enumerate(qs, start=1):
        full_name = (
            getattr(app.user, "full_name", "").strip() or
            app.user.get_username()
        )
        submitted_str = (
            app.submitted_at.strftime("%Y-%m-%d %H:%M")
            if app.submitted_at
            else ""
        )
        ws.append(
            [
                idx,
                full_name,
                app.user.mobile or "",
                app.application_no,
                app.course.course_name,
                app.course.degree,
                app.application_status,
                submitted_str,
            ]
        )

    for column_cells in ws.columns:
        max_len = 0
        col = column_cells[0].column
        for c in column_cells:
            val = str(c.value) if c.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    content_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response = HttpResponse(buffer.getvalue(), content_type=content_type)
    response["Content-Disposition"] = (
        'attachment; filename="pending_applications.xlsx"'
    )
    return response




#----------------------------------------------
#Contact Form 
#----------------------------------------------
def contact(request):
    if request.method == 'POST':
        form = UserContactForm(request.POST) 
        if form.is_valid():
            form.save()
            messages.success(request, "Thanks! Your message has been sent.")
            return redirect('contact')
    else:
        form = UserContactForm()
    return render(request, 'student/contact_form.html', {'form':form}) 


#Custom PasswordChange url redirect
class CustomPasswordChangeView(PasswordChangeView):
    success_url = reverse_lazy('student_dashboard') 
    
    def form_valid(self, form):
        messages.success(self.request, "Your password has been changed successfully.")
        return super().form_valid(form) 
    
#---------------------------------------
#Admin: Download PDF of Applicant Detail 
#---------------------------------------  

# views.py
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
from django.conf import settings
import os

def download_applicant_pdf(request, applicant_id):
    # Import your actual models
    from .models import Application, PersonalInfo, EducationalInfo, PaymentDetail
    
    # Get applicant data
    applicant = Application.objects.get(pk=applicant_id)
    per_info = PersonalInfo.objects.filter(user=applicant.user).first()
    edu_info = EducationalInfo.objects.filter(user=applicant.user)
    payment_info = PaymentDetail.objects.filter(application=applicant).first()
    
    # Create BytesIO buffer
    buffer = BytesIO()
    
    # Create PDF
    pdf = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.white,
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold',
        backColor=colors.HexColor('#0d6efd')
    )
    
    # Title
    title = Paragraph("Course Applicant Detail", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # 1. Application Information Section
    elements.append(Paragraph("Application Information", section_style))
    app_data = [
        ['Applicant:', applicant.user.get_full_name() or applicant.user.username],
        ['Email:', applicant.user.email or '—'],
        ['Mobile No.:', applicant.user.mobile or '—'],
        ['User Created at:', str(applicant.user.user_created_at) if applicant.user.user_created_at else '—'],
        ['Submitted:', applicant.submitted_at.strftime('%Y-%m-%d %H:%M:%S')],
    ]
    
    app_table = Table(app_data, colWidths=[2*inch, 4.5*inch])
    app_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8f9fa')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(app_table)
    elements.append(Spacer(1, 15))
    
    # 2. Personal Information Section with Profile Picture
    if per_info:
        elements.append(Paragraph("Personal Information", section_style))
        
        # Profile picture handling
        profile_pic = None
        if per_info.profile_pic:
            try:
                pic_path = per_info.profile_pic.path
                if os.path.exists(pic_path):
                    profile_pic = Image(pic_path, width=1.5*inch, height=1.5*inch)
            except:
                profile_pic = None
        
        personal_data = [
            ['Address:', per_info.address or '—'],
            ['Gender:', per_info.get_gender_display() if per_info.gender else '—'],
            ['Date of Birth:', per_info.dob.strftime('%Y-%m-%d') if per_info.dob else '—'],
            ['Father\'s Name:', per_info.father or '—'],
            ['Mother\'s Name:', per_info.mother or '—'],
            ['Grandfather\'s Name:', per_info.grandfather or '—'],
            ['Citizenship No.:', per_info.citizenship_no or '—'],
        ]
        
        # If profile pic exists, add it to the right side
        if profile_pic:
            per_table = Table(personal_data, colWidths=[2*inch, 3*inch])
            per_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8f9fa')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            combined_table = Table([[per_table, profile_pic]], colWidths=[5*inch, 1.8*inch])
            combined_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ]))
            elements.append(combined_table)
        else:
            per_table = Table(personal_data, colWidths=[2*inch, 4.5*inch])
            per_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8f9fa')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(per_table)
        
        elements.append(Spacer(1, 15))
    
    # 3. Educational Background Section
    elements.append(Paragraph("Educational Background", section_style))
    
    if edu_info.exists():
        edu_data = [['Level', 'Faculty', 'Course', 'University', 'College', 'Year', 'Grade/CGPA']]
        
        for edu in edu_info:
            edu_data.append([
                edu.level or '—',
                edu.faculty or '—',
                edu.course_name or '—',
                edu.university_name or '—',
                edu.college_name or '—',
                str(edu.passed_year) if edu.passed_year else '—',
                str(edu.grade_percent) if edu.grade_percent else '—',
            ])
        
        edu_table = Table(edu_data, colWidths=[0.9*inch, 1*inch, 1*inch, 1.3*inch, 1.3*inch, 0.7*inch, 0.8*inch])
        edu_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d1ecf1')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(edu_table)
    else:
        elements.append(Paragraph("No educational information available.", styles['Normal']))
    
    elements.append(Spacer(1, 15))
    
    # 4. Application Status Section
    elements.append(Paragraph("Application Status", section_style))
    
    status_data = [
        ['Degree:', applicant.course.degree],
        ['Course:', applicant.course.course_name],
        ['Application No.:', applicant.application_no],
        ['Submitted At:', str(applicant.submitted_at)],
        ['Status:', applicant.application_status],
    ]
    
    if applicant.application_status == 'approved' and applicant.approved_rejected_date:
        status_data.append(['Approved Date:', str(applicant.approved_rejected_date)])
    elif applicant.application_status == 'rejected':
        if applicant.reason_to_reject:
            status_data.append(['Rejection Reason:', applicant.reason_to_reject])
        if applicant.approved_rejected_date:
            status_data.append(['Rejection Date:', str(applicant.approved_rejected_date)])
    
    status_table = Table(status_data, colWidths=[2*inch, 4.5*inch])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#fff3cd')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(status_table)
    elements.append(Spacer(1, 15))
    
    # 5. Payment Details Section
    elements.append(Paragraph("Payment Details", section_style))
    
    if payment_info:
        payment_data = [
            ['Transaction ID:', payment_info.transaction_uuid],
            ['Status:', payment_info.status],
            ['Payment Method:', payment_info.payment_method],
            ['Amount:', str(payment_info.amount_paid)],
            ['Payment Date:', str(payment_info.payment_date)],
            ['Paid Course:', payment_info.application.course.course_name or '—'],
        ]
        
        payment_table = Table(payment_data, colWidths=[2*inch, 4.5*inch])
        payment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#d4edda')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(payment_table)
    else:
        elements.append(Paragraph("No payment has been recorded for this application.", styles['Normal']))
    
    # Build PDF
    pdf.build(elements)
    
    # Get PDF from buffer
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="applicant_{applicant.application_no}.pdf"'
    
    return response

